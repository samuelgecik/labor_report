from openpyxl import load_workbook

def _find_cell_by_text(sheet, search_texts):
    """
    Finds the first cell containing any of the provided search texts.
    
    Args:
        sheet: The openpyxl sheet object.
        search_texts: List of strings to search for.
    
    Returns:
        Tuple (row, col) 1-based if found, else None.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).lower()
                for text in search_texts:
                    if text.lower() in cell_str:
                        return (cell.row, cell.column)
    return None

def _get_real_cell_value(sheet, row, col):
    """
    Gets the value of a cell, handling merged cells correctly.
    
    Args:
        sheet: The openpyxl sheet object.
        row, col: 1-based indices.
    
    Returns:
        The cell value or the value from the top-left of merged range.
    """
    cell = sheet.cell(row, col)
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

def extract_data(
    file_path: str,
    column_mappings: dict,
    start_row_strategy=None,
    header_row_offset: int = 1,
    stop_condition: callable = None,
    extract_range: bool = False
) -> list[dict]:
    """
    Extracts and standardizes data from an Excel sheet.

    Args:
        file_path (str): The path to the Excel file.
        column_mappings (dict): A dictionary mapping standardized column
            names to a list of possible header texts to search for.
            Example: {'date': ['Dátum'], 'hours': ['odpracovaný čas', 'Počet odpracovaných hodín']}
        start_row_strategy (callable, optional): A function that takes the
            header row number as input and returns the starting row for data
            extraction. If None, the default strategy is to start from the
            row immediately after the header row.
        header_row_offset (int, optional): The number of rows to skip after
            the header row to find the start of the data. Defaults to 1.
        stop_condition (callable, optional): A function that takes the value
            of the cell in the first specified column (typically "date")
            as input and returns True if the extraction loop should terminate.
            Defaults to None.
        extract_range (bool, optional): If True, extracts all columns between
            the minimum and maximum column indices found via column_mappings (inclusive).
            The returned dictionary keys will be column indices as strings ('0', '1', ...).
            If False, behaves as before with standardized column names. Defaults to False.

    Returns:
        list[dict]: A list of dictionaries, where each dictionary
            represents a row of data with standardized column names if extract_range is False,
            otherwise with keys as column index strings ('0', '1', etc.).
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # Find column indices individually (handles multi-row headers)
    col_indices = {}
    for col_name, possible_headers in column_mappings.items():
        found = False
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_idx, col_idx)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    for header in possible_headers:
                        if header.lower() in cell_str:
                            col_indices[col_name] = col_idx
                            found = True
                            break
                if found:
                    break
            if found:
                break
        else:
            raise ValueError(f"Column for {col_name} not found")
    print(f"DEBUG: col_indices = {col_indices}")

    # Set header_row to the first row with any header text for compatibility
    all_header_texts = [text for texts in column_mappings.values() for text in texts]
    header_cell = _find_cell_by_text(sheet, all_header_texts)
    if header_cell:
        header_row, _ = header_cell
    else:
        # Fallback: use first row as header if nothing found
        header_row = 1

    # Calculate range if extract_range is True
    min_col = max_col = None
    if extract_range:
        if col_indices:
            min_col = min(col_indices.values())
            max_col = max(col_indices.values())
    print(f"DEBUG: min_col = {min_col}, max_col = {max_col}")
    
    # Determine start row
    if start_row_strategy:
        start_row = start_row_strategy(header_row)
    else:
        start_row = header_row + header_row_offset
    
    # Extract data
    data = []
    row = start_row
    date_column_idx = col_indices.get('date')
    while row <= sheet.max_row:
        row_data = {}
        valid_row = False
        if extract_range and min_col is not None:
            if row == start_row:
                print(f"DEBUG: First row extraction - column range: {min_col} to {max_col + 1}")
            for col_idx in range(min_col, max_col + 1):
                value = _get_real_cell_value(sheet, row, col_idx)
                row_data[str(col_idx - 1)] = value
                if value is not None:
                    valid_row = True
        else:
            for col_name, col_idx in col_indices.items():
                value = _get_real_cell_value(sheet, row, col_idx)
                row_data[col_name] = value
                if value is not None:
                    valid_row = True
        # Check stopping condition
        if extract_range:
            date_val = _get_real_cell_value(sheet, row, date_column_idx) if date_column_idx else None
        else:
            date_val = row_data.get('date')
        if not date_val or (stop_condition and stop_condition(date_val)):
            break
        if valid_row:
            data.append(row_data)
            if row == start_row:
                print(f"DEBUG: Extracted row_data keys: {list(row_data.keys())}")
        row += 1

    return data