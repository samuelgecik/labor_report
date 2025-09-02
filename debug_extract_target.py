import openpyxl
import csv
from openpyxl.utils import get_column_letter
from datetime import datetime, time, timedelta

def find_cell_by_text_partial(sheet, search_texts):
    """
    Find cell containing any of the search texts (partial match).
    Returns (row, col) or None.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for text in search_texts:
                    if text.lower() in cell.value.lower():
                        return cell.row, cell.column, get_column_letter(cell.column)
    return None

# Load the workbook
wb = openpyxl.load_workbook('ronec_vykaz.xlsx', data_only=True)
sheet = wb.active

# Handle merged ranges (if any)
merged_ranges = list(sheet.merged_cells)

def get_real_cell(row, col):
    for range_ in merged_ranges:
        if range_.min_row <= row <= range_.max_row and range_.min_col <= col <= range_.max_col:
            return sheet.cell(row=range_.min_row, column=range_.min_col)
    return sheet.cell(row=row, column=col)

# Dynamically locate 'Dátum' column
date_result = find_cell_by_text_partial(sheet, ['Dátum'])
if not date_result:
    print("Column 'Dátum' not found.")
    exit()

header_row = date_result[0]
date_col = date_result[1]
print(f"Found 'Dátum' at row {header_row}, column {get_column_letter(date_col)}")

# Dynamically locate 'Počet odpracovaných hodín*' column
hours_result = find_cell_by_text_partial(sheet, ['Počet odpracovaných hodín'])
if not hours_result:
    print("Column 'Počet odpracovaných hodín*' not found.")
    exit()

hours_col = hours_result[1]
print(f"Found 'Počet odpracovaných hodín*' at row {hours_result[0]}, column {get_column_letter(hours_col)}")

# Start extraction from row 26 (after header)
start_row = 26
print(f"Starting extraction from row {start_row}")

# Collect data
col_start = min(date_col, hours_col)
col_end = max(date_col, hours_col)
data = []
row = start_row

while True:
    date_val = get_real_cell(row, date_col).value
    if not date_val:  # Stop when date column becomes empty
        break
    row_data = []
    for col in range(col_start, col_end + 1):
        cell = get_real_cell(row, col)
        val = cell.value
        if val is None:
            row_data.append('')
        else:
            row_data.append(str(val))
    data.append(row_data)
    row += 1

print(f"Extracted {len(data)} rows")

# Export to CSV
with open('extracted_target.csv', 'w', newline='', encoding='utf-8') as csvfile:
    writer = csv.writer(csvfile)
    headers = [get_column_letter(col) for col in range(col_start, col_end + 1)]
    writer.writerow(headers)
    for row_data in data:
        writer.writerow(row_data)

print("Extraction completed. Data exported to 'extracted_target.csv'")