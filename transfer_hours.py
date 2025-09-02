import openpyxl
import datetime
from openpyxl.utils import get_column_letter

def find_cell_by_text_partial(sheet, search_texts):
    """
    Find cell containing any of the search texts (partial match).
    Returns (row, col) or None.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for text in search_texts:
                    if text.lower() in cell.value.lower():
                        return cell.row, cell.column, get_column_letter(cell.column)
    return None

def find_column_by_text(sheet, search_text):
    """
    Find exact column by text.
    Returns col number or None.
    """
    result = find_cell_by_text_partial(sheet, [search_text])
    return result[1] if result else None

def process_hours(h_val):
    if isinstance(h_val, datetime.time):
        return h_val.hour + h_val.minute / 60.0 + h_val.second / 3600.0
    elif isinstance(h_val, str) and h_val == '-':
        return 0.0
    else:
        return h_val

# Load source workbook
source_wb = openpyxl.load_workbook('ronec_dochadzka.xlsx')
source_sheet = source_wb.active

# Load target workbook
target_wb = openpyxl.load_workbook('ronec_vykaz.xlsx')
target_sheet = target_wb.active

# Find dynamic columns
source_date_col = find_column_by_text(source_sheet, 'Dátum') or 2  # Default column B
source_hours_col = find_column_by_text(source_sheet, 'odpracovaný čas') or 8  # Default column H
prichod_result = find_cell_by_text_partial(source_sheet, ['Príchod', 'Príchd'])
prichod_col = prichod_result[1] if prichod_result else None
target_desc_col = find_column_by_text(target_sheet, 'Detailný popis činností vykonávaných na základe Zmluvy o PPM a popis zrealizovaných výstupov')
target_date_col = find_column_by_text(target_sheet, 'Dátum') or 1  # Assuming column A
target_hours_col = find_column_by_text(target_sheet, 'Počet odpracovaných hodín*') or 9  # Assuming column I


# Collect merged ranges
merged_ranges = list(target_sheet.merged_cells)

def get_real_cell(row, col):
    for range_ in merged_ranges:
        if range_.min_row <= row <= range_.max_row and range_.min_col <= col <= range_.max_col:
            return target_sheet.cell(row=range_.min_row, column=range_.min_col)
    return target_sheet.cell(row=row, column=col)

# Collect July date and hours data from source
data = []
row = 7  # Data starts below headers (after row 6)

while True:
    date = source_sheet.cell(row=row, column=source_date_col).value
    hours = source_sheet.cell(row=row, column=source_hours_col).value
    if not date or not hours:
        break
    # Filter for July dates (month 7)
    if isinstance(date, datetime.date) and date.month == 7:
        date_day = str(date.day) + '.'
        data.append((date_day, hours))
    row += 1

# Collect Dovolenka dates from 'Príchod' column
dovolenka_dates = []
if prichod_col:
    row = 7
    max_row = source_sheet.max_row
    while row <= max_row:
        prichod_val = source_sheet.cell(row=row, column=prichod_col).value
        date_val = source_sheet.cell(row=row, column=source_date_col).value
        if isinstance(prichod_val, str) and prichod_val.strip() == 'Dovolenka' and isinstance(date_val, datetime.date) and date_val.month == 7:
            dovolenka_dates.append(str(date_val.day) + '.')
        row += 1

# Transfer data to target
for i, (date_val, hours_val) in enumerate(data):
    target_row = 26 + i
    get_real_cell(target_row, target_date_col).value = date_val
    get_real_cell(target_row, target_hours_col).value = hours_val

# Transfer Dovolenka to target description column
if target_desc_col and dovolenka_dates:
    # Build date to row mapping in target
    target_date_rows = {}
    for row in range(26, target_sheet.max_row + 1):  # Assuming data starts at row 26
        date_cell = get_real_cell(row, target_date_col)
        if date_cell.value:
            target_date_rows[str(date_cell.value).strip()] = row

    # Set Dovolenka for matching dates
    for dov_date in dovolenka_dates:
        if dov_date in target_date_rows:
            desc_cell = get_real_cell(target_date_rows[dov_date], target_desc_col)
            if desc_cell:
                desc_cell.value = 'Dovolenka'

# Save updated target file
target_wb.save('ronec_vykaz_updated.xlsx')

# Signal completion
print(f"Transfer completed. Total records with hours: {len(data)}, Dovolenka entries: {len(dovolenka_dates)}")