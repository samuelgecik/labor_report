import openpyxl
from datetime import datetime

# Load source workbook
source_wb = openpyxl.load_workbook('ronec_dochadzka.xlsx')
source_sheet = source_wb.active

print("Debugging data from ronec_dochadzka.xlsx")
print("Source: Dates from Row 6 Column B (Column 2), Hours from Row 7 Column H (Column 8)")
print()

# Print header context (rows around data start)
for row in range(4, 8):  # Rows 4-7 to see headers and first data
    date_val = source_sheet.cell(row=row, column=2).value
    hours_val = source_sheet.cell(row=row, column=8).value if row >= 7 else None
    print(f"Row {row}: Date (B) = {date_val} (type: {type(date_val)}), Hours (H) = {hours_val} (type: {type(hours_val) if hours_val is not None else 'N/A'})")

print()
print("Full data check (rows 6-20 for dates, rows 7-20 for hours):")
for row in range(6, 21):  # Check rows 6-20
    date_val = source_sheet.cell(row=row, column=2).value
    hours_val = source_sheet.cell(row=row, column=8).value if row >= 7 else None

    # Data type and July check for dates
    date_info = ""
    if date_val:
        if isinstance(date_val, str):
            date_info = f" [is str, contains '.7.': {'.7.' in date_val}]"
        elif isinstance(date_val, datetime):
            month = date_val.month
            date_info = f" [is datetime, month: {month}, is July: {month == 7}]"

    # Data type for hours
    hours_info = ""
    if hours_val:
        hours_info = f" (type: {type(hours_val)})"

    print(f"Row {row}: Date (B) = {date_val}{date_info}, Hours (H) = {hours_val}{hours_info}")