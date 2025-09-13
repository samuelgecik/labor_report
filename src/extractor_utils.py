"""
Reusable utility functions for extracting data from Excel workbooks.

This module now contains both:
 - Low-level extract_data implementation (migrated from src/extract.py)
 - Higher-level helpers (strategy registry, multi-sheet extraction, CSV saving)
 - Workbook handling utilities (opening, backup creation)
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Any, Union, Tuple
from openpyxl import load_workbook


# Strategy Registry for callable functions
STRATEGY_REGISTRY = {
    "source": {
            "column_indices": [1, 2, 3, 4, 5, 6, 7],
            "header_text": "Dátum",
            "header_row_offset": 2,
            "start_row_strategy": None,
            "stop_condition": None
        
    },
    "target": {
            "column_indices": [1, 2, 3, 4, [5, 6, 7, 8], 9, 10, 11, 12, 13, 14],
            "start_row_strategy": lambda header_row: 26,  # fixed_26
            "stop_condition": lambda row_data: len(row_data) > 4 and row_data[4] and "Spolu:" in str(row_data[4]),  # stop_for_spolu
            "header_text": None,
            "header_row_offset": 1
    },
}


# -------------------------------
# Low-level extraction primitives
# -------------------------------
def _find_cell_by_text(sheet, search_texts: List[str]):
    """Find first cell containing any provided search strings; returns (row, col) 1-based or None."""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).lower()
                for text in search_texts:
                    if text.lower() in cell_str:
                        return (cell.row, cell.column)
    return None


def _get_real_cell_value(sheet, row: int, col: int):
    """Return cell value, following merged ranges to the top-left cell when applicable."""
    cell = sheet.cell(row, col)
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def extract_data(
    file_path: str,
    column_indices: List[Union[int, List[int]]],
    header_text: str | None = None,
    start_row_strategy=None,
    header_row_offset: int = 1,
    stop_condition: callable = None,
    sheet_name: str = None,
) -> List[List[Any]]:
    """Extract values from selected columns of an Excel sheet with optional header/start/stop logic."""
    wb = load_workbook(file_path, data_only=True)
    if sheet_name:
        try:
            sheet = wb[sheet_name]
        except KeyError:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}. Available sheets: {list(wb.sheetnames)}")
    else:
        sheet = wb.active

    # Determine header position and starting column
    if header_text:
        pos = _find_cell_by_text(sheet, [header_text])
        if pos:
            header_row = pos[0]
            starting_col = pos[1]
        else:
            header_row = 1
            starting_col = 1
    else:
        header_row = 1
        starting_col = 1

    # Determine start row
    start_row = start_row_strategy(header_row) if start_row_strategy else (header_row + header_row_offset)

    # Row extraction loop
    data: List[List[Any]] = []
    row = start_row
    while row <= sheet.max_row:
        row_data: List[Any] = []
        for col_idx in column_indices:
            if isinstance(col_idx, int):
                actual_col = starting_col + (col_idx - 1)
                value = _get_real_cell_value(sheet, row, actual_col)
            elif isinstance(col_idx, list):
                value = None
                dovolenka_found = False
                for inner_col_idx in col_idx:
                    actual_inner_col = starting_col + (inner_col_idx - 1)
                    cell_value = _get_real_cell_value(sheet, row, actual_inner_col)
                    if cell_value is not None and "Dovolenka" in str(cell_value):
                        value = cell_value
                        dovolenka_found = True
                        break
                if not dovolenka_found:
                    for inner_col_idx in col_idx:
                        actual_inner_col = starting_col + (inner_col_idx - 1)
                        cell_value = _get_real_cell_value(sheet, row, actual_inner_col)
                        if cell_value is not None and str(cell_value).strip():
                            value = cell_value
                            break
            else:
                value = None
            row_data.append(value)

        # Evaluate stop condition before storing row
        if stop_condition and stop_condition(row_data):
            break

        # If row contains any non-None value, keep it; else stop on first empty row
        if any(v is not None for v in row_data):
            data.append(row_data)
        else:
            break
        row += 1

    wb.close()
    return data


def extract_from_workbook(config: Dict[str, Any]) -> Dict[str, List[List[Any]]]:
    """
    Extracts data from multiple sheets in a workbook based on configuration or strategy.

    Args:
        config: Dictionary containing extraction configuration with keys:
            - file_path: str - Path to the Excel file
            - sheets: str | list - Either "__ALL__" for all sheets,
                                    or a list of specific sheet names
            - strategy: str (optional) - Strategy name from STRATEGY_REGISTRY
            - column_indices: List of column specifications (if not using strategy)
            - header_text: str (optional) - Text to search for header row
            - header_row_offset: int (optional) - Offset from header row to start data
            - start_row_strategy: str (optional) - Key into STRATEGY_REGISTRY
            - stop_condition: str (optional) - Key into STRATEGY_REGISTRY

    Returns:
        Dict[str, List[List]]: Dictionary where keys are sheet names and values
                               are the extracted data for each sheet as list of lists
    """
    results = {}

    # Load the workbook to get sheet names
    wb = load_workbook(config['file_path'], read_only=True)

    # Determine which sheets to process
    sheets_to_process = []
    if config.get('sheets') == "__ALL__":
        sheets_to_process = wb.sheetnames
    elif isinstance(config['sheets'], list):
        sheets_to_process = config['sheets']
    else:
        # Single sheet name
        sheets_to_process = [config['sheets']]

    # Get strategy configuration if specified
    strategy_name = config.get('strategy')
    if strategy_name:
        # Look for strategy in source or target registries
        strategy_config = STRATEGY_REGISTRY.get('source', {}).get(strategy_name) or \
                         STRATEGY_REGISTRY.get('target', {}).get(strategy_name)
        if strategy_config and isinstance(strategy_config, dict):
            # Merge strategy defaults with config overrides
            merged_config = strategy_config.copy()
            merged_config.update({k: v for k, v in config.items() 
                                if k not in ['strategy', 'file_path', 'sheets']})
        else:
            merged_config = config
    else:
        merged_config = config

    # Process each sheet
    for sheet_name in sheets_to_process:
        # Resolve strategy functions from registry (legacy support)
        start_strategy_key = merged_config.get('start_row_strategy')
        stop_condition_key = merged_config.get('stop_condition')
        
        start_strategy = None
        stop_condition = None
        
        if isinstance(start_strategy_key, str):
            start_strategy = STRATEGY_REGISTRY.get('target', {}).get(start_strategy_key)
        elif callable(start_strategy_key):
            start_strategy = start_strategy_key
            
        if isinstance(stop_condition_key, str):
            stop_condition = STRATEGY_REGISTRY.get('target', {}).get(stop_condition_key)
        elif callable(stop_condition_key):
            stop_condition = stop_condition_key

        # Prepare the arguments for extract_data
        extract_args = {
            'file_path': config['file_path'],
            'column_indices': merged_config['column_indices'],
            'sheet_name': sheet_name
        }

        # Add optional parameters if provided
        if merged_config.get('header_text') is not None:
            extract_args['header_text'] = merged_config['header_text']
        if merged_config.get('header_row_offset') is not None:
            extract_args['header_row_offset'] = merged_config['header_row_offset']
        if start_strategy:
            extract_args['start_row_strategy'] = start_strategy
        if stop_condition:
            extract_args['stop_condition'] = stop_condition

        # Extract data from this sheet
        try:
            data = extract_data(**extract_args)
            results[sheet_name] = data
        except Exception as e:
            logging.error(f"Failed to extract data from sheet '{sheet_name}': {e}")
            results[sheet_name] = []

    wb.close()  # Close the readonly workbook
    return results


def extract_whole_workbook(file_path: str, strategy: str = "perry_soft") -> Dict[str, List[List[Any]]]:
    """
    Extract data from all sheets in a workbook using a predefined strategy.
    
    Args:
        file_path: Path to the Excel file
        strategy: Strategy name from STRATEGY_REGISTRY ('perry_soft', 'attendance_default', 'vykaz_template')
    
    Returns:
        Dict[str, List[List]]: Dictionary where keys are sheet names and values are extracted data
    """
    config = {
        'file_path': file_path,
        'sheets': "__ALL__",
        'strategy': strategy
    }
    
    return extract_from_workbook(config)


def save_extraction_results(results: Dict[str, List[List[Any]]], config: Dict[str, Any]) -> None:
    """
    Saves extraction results to CSV files.

    Args:
        results: Dictionary of sheet names to extracted data
        config: Configuration dictionary containing 'output_prefix' for naming
    """
    import csv
    import os

    output_prefix = config.get('output_prefix', 'extracted_data')
    headers = config.get('headers', [])

    for sheet_name, data in results.items():
        # Create a safe filename from sheet name
        safe_sheet_name = sheet_name.replace(' ', '_').replace('/', '_')
        filename = f"{output_prefix}_{safe_sheet_name}.csv"
        output_path = os.path.join('data', 'output', filename)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        with open(output_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            if headers:
                writer.writerow(headers)
            writer.writerows(data)


def open_workbooks(source_excel: str, target_excel: str, backup_dir: str, dry_run: bool) -> Tuple[Any, Any, str | None]:
    """Open source (read-only) and target (write) workbooks.

    Creates timestamped backup of target (unless dry_run) under backup_dir.

    Returns (source_wb, target_wb, backup_path)
    """
    if not os.path.exists(source_excel):
        raise SystemExit(f"Source workbook not found: {source_excel}")
    if not os.path.exists(target_excel):
        raise SystemExit(f"Target workbook not found: {target_excel}")

    source_wb = load_workbook(source_excel, read_only=True, data_only=True)
    target_wb = load_workbook(target_excel)
    backup_path = None

    if not dry_run:
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_filename)
        # Simple copy by saving a duplicate workbook object
        target_wb.save(backup_path)
        logging.info(f"Created backup of target workbook: {backup_path}")
    else:
        logging.info("Dry-run: skipping target backup creation")

    return source_wb, target_wb, backup_path