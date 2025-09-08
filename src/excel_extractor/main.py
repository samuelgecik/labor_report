from openpyxl import load_workbook
from typing import List, Union

def _find_cell_by_text(sheet, search_texts):
    """
    Finds the first cell containing any of the provided search texts.
    
    Args:
        sheet: The openpyxl sheet object.
        search_texts: List of strings to search for.
    
    Returns:
        Tuple (row, col) 1-based if found, else None.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).lower()
                for text in search_texts:
                    if text.lower() in cell_str:
                        return (cell.row, cell.column)
    return None

def _get_real_cell_value(sheet, row, col):
    """
    Gets the value of a cell, handling merged cells correctly.
    
    Args:
        sheet: The openpyxl sheet object.
        row, col: 1-based indices.
    
    Returns:
        The cell value or the value from the top-left of merged range.
    """
    cell = sheet.cell(row, col)
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value
    

def extract_data(
    file_path: str,
    column_indices: List[Union[int, List[int]]],
    header_text: str | None = None,
    start_row_strategy=None,
    header_row_offset: int = 1,
    stop_condition: callable = None
) -> list[list]:
    """
    Extracts data from specified columns in an Excel sheet.

    Args:
        file_path (str): The path to the Excel file.
        column_indices (List[Union[int, List[int]]]): List where each item can be either:
            - int: A 1-based column number for direct 1-to-1 mapping (e.g., 1 for column A)
            - List[int]: A list of 1-based column numbers for coalesce operation,
              where the first non-empty value across those columns will be used
              (e.g., [5, 6, 7, 8] to pick the first non-empty value from E, F, G, H)
        header_text (str, optional): Text to search for to locate the header row.
            If provided, the function finds the row containing this text and uses it as the header row.
            If not provided, defaults to row 1.
        start_row_strategy (callable, optional): A function that takes the
            header row number as input and returns the starting row for data
            extraction. If None, the default strategy is to start from the
            row immediately after the header row.
        header_row_offset (int, optional): The number of rows to skip after
            the header row to find the start of the data. Defaults to 1.
        stop_condition (callable, optional): A function that takes the row data (list) as input and returns True to stop extraction. Defaults to None.

    Returns:
        list[list]: A list of lists where each inner list represents a row of
            data values from the specified columns.
    """
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Determine header_row
    if header_text:
        pos = _find_cell_by_text(sheet, [header_text])
        if pos:
            header_row = pos[0]
            starting_col = pos[1]
        else:
            header_row = 1  # fallback if text not found
            starting_col = 1
    else:
        header_row = 1
        starting_col = 1

    # Determine start row
    if start_row_strategy:
        start_row = start_row_strategy(header_row)
    else:
        start_row = header_row + header_row_offset

    # Extract data
    data = []
    row = start_row
    while row <= sheet.max_row:
        row_data = []
        for col_idx in column_indices:
            if isinstance(col_idx, int):
                # Direct 1-to-1 column mapping
                actual_col = starting_col + (col_idx - 1)
                value = _get_real_cell_value(sheet, row, actual_col)
            elif isinstance(col_idx, list):
                # Coalesce operation: prioritize "Dovolenka", else first non-empty
                value = None
                dovolenka_found = False
                for inner_col_idx in col_idx:
                    actual_inner_col = starting_col + (inner_col_idx - 1)
                    cell_value = _get_real_cell_value(sheet, row, actual_inner_col)
                    if cell_value is not None and "Dovolenka" in str(cell_value):
                        value = cell_value
                        dovolenka_found = True
                        break
                if not dovolenka_found:
                    for inner_col_idx in col_idx:
                        actual_inner_col = starting_col + (inner_col_idx - 1)
                        cell_value = _get_real_cell_value(sheet, row, actual_inner_col)
                        if cell_value is not None and str(cell_value).strip():
                            value = cell_value
                            break
            else:
                # Fallback for unexpected types
                value = None
            row_data.append(value)

        # Check stopping condition before appending
        if stop_condition and stop_condition(row_data):
            break

        # Check if row has any non-None data
        if any(v is not None for v in row_data):
            data.append(row_data)
        else:
            # If the entire row is None, stop extraction (empty row)
            break

        row += 1

    return data