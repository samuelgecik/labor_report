import argparse
import calendar
import shutil
import os
import logging
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Any, Optional
import pandas as pd
from openpyxl import load_workbook

from src.extractor_utils import extract_from_workbook, open_workbooks
from src import sheet_mapper

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments for updating labor report workbook."""
    parser = argparse.ArgumentParser(description="Update labor report workbook")
    parser.add_argument("--source-excel", required=False, default=None,
                       help="Path to source attendance/work log Excel file")
    parser.add_argument("--vacations-json", required=False, default=None,
                       help="Path to vacations JSON file (used when no source-excel)")
    parser.add_argument("--target-excel", required=True, 
                       help="Path to target labor report Excel file")
    parser.add_argument("--month", type=str, default=None, 
                       help="Month name in Slovak (e.g., 'júl')")
    parser.add_argument("--year", type=int, default=None, 
                       help="Year (e.g., 2025)")
    parser.add_argument("--activity-text", type=str, default=None,
                       help="Activity description text to use")
    parser.add_argument("--work-location", type=str, default="Bratislava",
                       help="Work location (default: Bratislava)")
    parser.add_argument("--dry-run", action="store_true", 
                       help="Run validation only; do not write changes")
    parser.add_argument("--output-dir", default="data/output", 
                           help="Directory to save outputs")
    parser.add_argument("--no-clean-target", action="store_true", default=False,
                       help="Skip removing unmatched target sheets")
    parser.add_argument("--no-sort-target", action="store_true", default=False,
                       help="Skip sorting target sheets based on source sheet order")
    return parser.parse_args()


def extract_source_data(source_excel: str, sheet_name: str = None) -> pd.DataFrame:
    """Extract source data from Excel file using extractor_utils."""
    # Use the source strategy directly from STRATEGY_REGISTRY
    from src.extractor_utils import STRATEGY_REGISTRY
    
    source_strategy = STRATEGY_REGISTRY["source"]
    config = {
        'file_path': source_excel,
        'sheets': [sheet_name] if sheet_name else "__ALL__",  # Extract from specific sheet or all sheets
        'column_indices': source_strategy["column_indices"],
        'header_text': source_strategy["header_text"],
        'header_row_offset': source_strategy["header_row_offset"],
        'start_row_strategy': source_strategy["start_row_strategy"],
        'stop_condition': source_strategy["stop_condition"]
    }
    
    results = extract_from_workbook(config)
    
    # For now, use the first sheet's data (or specified sheet)
    if not results:
        raise ValueError(f"No data extracted from {source_excel}")
    
    # Get the target sheet's data
    if sheet_name and sheet_name in results:
        data = results[sheet_name]
    else:
        sheet_name = list(results.keys())[0]
        data = results[sheet_name]
    
    # Convert to DataFrame with expected column names
    columns = ['Datum', 'Dochadzka_Prichod', 'Dochadzka_Odchod', 'Prestavka_min', 
               'Prerusenie_Odchod', 'Prerusenie_Prichod', 'Skutocny_Odpracovany_Cas']
    
    df = pd.DataFrame(data, columns=columns)
    
    # Clean and process data
    df['Skutocny_Odpracovany_Cas'] = df['Skutocny_Odpracovany_Cas'].astype(str).str.strip().replace(' -', '-', regex=False)
    df['Prestavka_min'] = df['Prestavka_min'].astype(str).str.strip().replace(' -', '-', regex=False)
    
    # Parse dates
    df['Datum'] = pd.to_datetime(df['Datum'], errors='coerce')
    
    # Clean NaN values
    numeric_cols = df.select_dtypes(include=['number']).columns
    df[numeric_cols] = df[numeric_cols].fillna('-')
    
    return df


def source_to_target(df_source: pd.DataFrame, activity_text: str, work_location: str) -> pd.DataFrame:
    """Transform source data to target format."""
    cols = ['Datum', 'Cas_Vykonu_Od', 'Cas_Vykonu_Do', 'Prestavka_Trvanie', 
            'Popis_Cinnosti', 'Pocet_Odpracovanych_Hodin', 'Miesto_Vykonu', 
            'PH_Projekt_POO', 'PH_Riesenie_POO', 'PH_Mimo_Projekt_POO', 'SPOLU']
    
    df_target = pd.DataFrame(columns=cols, index=range(31))
    
    # Extract day numbers for all 31 days
    df_target['Datum'] = [str(i + 1) + '.' for i in range(31)]
    
    def get_prestavka(row):
        p_min = row['Prestavka_min']
        if p_min == '-' or pd.isna(p_min):
            return '00:00:00'
        if isinstance(p_min, str) and p_min.isdigit():
            mins = int(p_min)
        elif isinstance(p_min, (int, float)):
            mins = p_min
        else:
            return '00:00:00'
        try:
            td = timedelta(minutes=mins)
            hours = td.seconds // 3600
            mins_part = (td.seconds % 3600) // 60
            return f"{hours:02}:{mins_part:02}:00"
        except:
            return '00:00:00'
    
    # Default activity text if none provided
    if not activity_text:
        activity_text = "Pracovná činnosť"
    
    def _sanitize_time(value):
        """Ensure time values are in HH:MM:SS string format. Return empty string for invalid values."""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ''
        # If it's a set containing a single string, unwrap it
        if isinstance(value, set) and len(value) == 1:
            value = next(iter(value))
        # If it's a number (minutes), convert to HH:MM:SS
        if isinstance(value, (int, float)) and not pd.isna(value):
            try:
                mins = int(value)
                td = timedelta(minutes=mins)
                hours = td.seconds // 3600
                mins_part = (td.seconds % 3600) // 60
                return f"{hours:02}:{mins_part:02}:00"
            except Exception:
                return ''
        # If it's already a string, try to normalize common cases
        if isinstance(value, str):
            v = value.strip()
            if v == '-' or v == '':
                return ''
            # Common already-HH:MM:SS
            if ':' in v:
                parts = v.split(':')
                if len(parts) == 2:
                    # mm:ss or hh:mm -> make hh:mm:00
                    return f"{int(parts[0]):02}:{int(parts[1]):02}:00"
                if len(parts) == 3:
                    try:
                        h, m, s = map(int, parts)
                        return f"{h:02}:{m:02}:{s:02}"
                    except Exception:
                        return v
            # Try parse as integer minutes string
            if v.isdigit():
                try:
                    mins = int(v)
                    td = timedelta(minutes=mins)
                    hours = td.seconds // 3600
                    mins_part = (td.seconds % 3600) // 60
                    return f"{hours:02}:{mins_part:02}:00"
                except Exception:
                    return ''
            return v

    for i in range(31):
        if i < len(df_source):
            row = df_source.iloc[i]
            dochadzka = row['Dochadzka_Prichod']
        else:
            dochadzka = '-'
            row = None
        
        # Compact template-based approach
        templates = {
            'vacation': {'Popis_Cinnosti': 'DOVOLENKA', 'Pocet_Odpracovanych_Hodin': row['Skutocny_Odpracovany_Cas'] if row is not None else '00:00:00', 'SPOLU': row['Skutocny_Odpracovany_Cas'] if row is not None else '00:00:00'},
            'absent': {'Prestavka_Trvanie': '00:00:00'},
            'weekend': {'Prestavka_Trvanie': '00:00:00'}
        }
        
        # Set defaults for all non-work days
        zero_fields = ['PH_Projekt_POO', 'PH_Riesenie_POO', 'PH_Mimo_Projekt_POO']
        empty_fields = ['Cas_Vykonu_Od', 'Cas_Vykonu_Do', 'Miesto_Vykonu', 'Popis_Cinnosti']
        
        # Determine day type
        if dochadzka == 'Dovolenka':
            day_type = 'vacation'
        elif (dochadzka == '-' or pd.isna(dochadzka) or 
              (row is not None and not pd.isna(row['Datum']) and 
               all(pd.isna(row[col]) or str(row[col]).strip() == '-' 
                   for col in ['Dochadzka_Prichod', 'Dochadzka_Odchod', 'Prestavka_min', 
                               'Prerusenie_Odchod', 'Prerusenie_Prichod', 'Skutocny_Odpracovany_Cas']))):
            day_type = 'weekend' if (row is not None and not pd.isna(row['Datum'])) else 'absent'
        else:
            day_type = 'work'
        
        if day_type != 'work':
            # Apply non-work template
            for field in zero_fields:
                df_target.loc[i, field] = '00:00:00'
            for field in empty_fields:
                df_target.loc[i, field] = ''
            df_target.loc[i, 'Pocet_Odpracovanych_Hodin'] = '00:00:00'
            df_target.loc[i, 'SPOLU'] = '00:00:00'
            # Apply specific template overrides
            for field, value in templates.get(day_type, {}).items():
                # sanitize when setting template values
                if field in ('Pocet_Odpracovanych_Hodin', 'SPOLU', 'Prestavka_Trvanie'):
                    df_target.loc[i, field] = _sanitize_time(value)
                else:
                    df_target.loc[i, field] = value
            logging.info(f"Applied {day_type} template to row {i}")
        else:
            # Work day
            worked_hours = row['Skutocny_Odpracovany_Cas']
            df_target.loc[i, 'Cas_Vykonu_Od'] = row['Dochadzka_Prichod']
            df_target.loc[i, 'Cas_Vykonu_Do'] = row['Dochadzka_Odchod']
            df_target.loc[i, 'Prestavka_Trvanie'] = _sanitize_time(get_prestavka(row))
            df_target.loc[i, 'Popis_Cinnosti'] = activity_text
            df_target.loc[i, 'Pocet_Odpracovanych_Hodin'] = worked_hours
            df_target.loc[i, 'Miesto_Vykonu'] = work_location
            for field in zero_fields:
                df_target.loc[i, field] = '00:00:00'
            df_target.loc[i, 'SPOLU'] = worked_hours
            logging.info(f"Applied work template to row {i}")
    
    return df_target


def _easter_sunday(year: int) -> date:
    """Gregorian Easter Sunday (Anonymous/Meeus algorithm)."""
    a = year % 19
    b, c = divmod(year, 100)
    d, e = divmod(b, 4)
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i, k = divmod(c, 4)
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def slovak_days_of_rest(year: int, month: int) -> set:
    """Return the day-numbers in (year, month) that are Slovak paid days of rest
    (dni pracovného pokoja) — i.e. non-work days for timesheet generation.

    Reflects the consolidation-package changes (grounded; see project memory
    'slovak-days-of-rest'):
      - Deň Ústavy (1 Sep) and Vznik ČSR (28 Oct): NOT a day of rest since 2024.
      - Deň boja za slobodu (17 Nov): NOT a day of rest since 2025.
      - Deň víťazstva (8 May) and Sedembolestná P. Márie (15 Sep): suspended for 2026.
    Movable feasts (Veľký piatok, Veľkonočný pondelok) are derived from Easter.
    """
    fixed = {
        (1, 1): True,             # Deň vzniku Slovenskej republiky
        (1, 6): True,             # Zjavenie Pána (Traja králi)
        (5, 1): True,             # Sviatok práce
        (5, 8): year != 2026,     # Deň víťazstva nad fašizmom — suspended for 2026
        (7, 5): True,             # Sviatok sv. Cyrila a Metoda
        (8, 29): True,            # Výročie SNP
        (9, 1): year < 2024,      # Deň Ústavy SR — working day since 2024
        (9, 15): year != 2026,    # Sedembolestná Panna Mária — suspended for 2026
        (10, 28): year < 2024,    # Vznik samostatného česko-slovenského štátu — working since 2024
        (11, 1): True,            # Sviatok všetkých svätých
        (11, 17): year < 2025,    # Deň boja za slobodu a demokraciu — working day since 2025
        (12, 24): True,           # Štedrý deň
        (12, 25): True,           # Prvý sviatok vianočný
        (12, 26): True,           # Druhý sviatok vianočný
    }
    days = {d for (mo, d), is_rest in fixed.items() if mo == month and is_rest}

    easter = _easter_sunday(year)
    for offset in (-2, 1):  # Veľký piatok (Easter-2), Veľkonočný pondelok (Easter+1)
        feast = easter + timedelta(days=offset)
        if feast.month == month and feast.year == year:
            days.add(feast.day)
    return days


def generate_contractor_data(year: int, month: int, activity_text: str, work_location: str) -> pd.DataFrame:
    """Generate a target DataFrame for a contractor with standard 8-hour shifts on business days.

    Args:
        year: Calendar year
        month: Month number (1-12)
        activity_text: Activity description text
        work_location: Work location string

    Returns:
        DataFrame with 31 rows in the same format as source_to_target output.
    """
    cols = ['Datum', 'Cas_Vykonu_Od', 'Cas_Vykonu_Do', 'Prestavka_Trvanie',
            'Popis_Cinnosti', 'Pocet_Odpracovanych_Hodin', 'Miesto_Vykonu',
            'PH_Projekt_POO', 'PH_Riesenie_POO', 'PH_Mimo_Projekt_POO', 'SPOLU']

    df = pd.DataFrame(columns=cols, index=range(31))
    df['Datum'] = [str(i + 1) + '.' for i in range(31)]

    if not activity_text:
        activity_text = "Pracovná činnosť"

    days_in_month = calendar.monthrange(year, month)[1]
    rest_days = slovak_days_of_rest(year, month)
    zero_fields = ['PH_Projekt_POO', 'PH_Riesenie_POO', 'PH_Mimo_Projekt_POO']

    for i in range(31):
        day_num = i + 1
        for field in zero_fields:
            df.loc[i, field] = '00:00:00'

        if day_num > days_in_month:
            # Day doesn't exist in this month — treat as absent
            df.loc[i, 'Cas_Vykonu_Od'] = ''
            df.loc[i, 'Cas_Vykonu_Do'] = ''
            df.loc[i, 'Prestavka_Trvanie'] = '00:00:00'
            df.loc[i, 'Popis_Cinnosti'] = ''
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '00:00:00'
            df.loc[i, 'Miesto_Vykonu'] = ''
            df.loc[i, 'SPOLU'] = '00:00:00'
            logging.info(f"Applied absent template to row {i} (day {day_num} beyond month)")
            continue

        weekday = calendar.weekday(year, month, day_num)  # 0=Mon, 6=Sun

        if weekday >= 5 or day_num in rest_days:
            # Weekend or Slovak public holiday (deň pracovného pokoja) — non-work
            df.loc[i, 'Cas_Vykonu_Od'] = ''
            df.loc[i, 'Cas_Vykonu_Do'] = ''
            df.loc[i, 'Prestavka_Trvanie'] = '00:00:00'
            df.loc[i, 'Popis_Cinnosti'] = ''
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '00:00:00'
            df.loc[i, 'Miesto_Vykonu'] = ''
            df.loc[i, 'SPOLU'] = '00:00:00'
            label = 'weekend' if weekday >= 5 else 'holiday'
            logging.info(f"Applied {label} template to row {i}")
        else:
            # Business day — standard 8-hour shift
            df.loc[i, 'Cas_Vykonu_Od'] = '09:00:00'
            df.loc[i, 'Cas_Vykonu_Do'] = '17:30:00'
            df.loc[i, 'Prestavka_Trvanie'] = '00:30:00'
            df.loc[i, 'Popis_Cinnosti'] = activity_text
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '08:00:00'
            df.loc[i, 'Miesto_Vykonu'] = work_location
            df.loc[i, 'SPOLU'] = '08:00:00'
            logging.info(f"Applied contractor work template to row {i}")

    return df


def generate_data_with_vacations(year: int, month: int, activity_text: str,
                                  work_location: str, vacation_days: List = None) -> pd.DataFrame:
    """Generate target DataFrame with standard 8h shifts, applying vacation overrides.

    Args:
        year: Calendar year
        month: Month number (1-12)
        activity_text: Activity description text
        work_location: Work location string
        vacation_days: List of vacation entries from JSON — integers for full days,
                       dicts with 'day' and 'half' keys for half-day vacations.

    Returns:
        DataFrame with 31 rows in the same format as source_to_target output.
    """
    # Start with standard contractor data (8h on business days)
    df = generate_contractor_data(year, month, activity_text, work_location)

    if not vacation_days:
        return df

    # Build lookup: day_num -> vacation type
    # 'full' for full-day, 'morning'/'afternoon' for half-day (which half is vacation)
    vac_lookup = {}
    for entry in vacation_days:
        if isinstance(entry, int):
            vac_lookup[entry] = 'full'
        elif isinstance(entry, dict):
            vac_lookup[entry['day']] = entry['half']

    days_in_month = calendar.monthrange(year, month)[1]
    rest_days = slovak_days_of_rest(year, month)

    for day_num, vac_type in vac_lookup.items():
        if day_num > days_in_month:
            continue
        if day_num in rest_days:
            # A public holiday is a day off already — it is never recorded as vacation.
            logging.info(f"Skipping vacation on day {day_num} (Slovak public holiday)")
            continue
        i = day_num - 1  # row index

        if vac_type == 'full':
            # Full-day vacation
            df.loc[i, 'Cas_Vykonu_Od'] = ''
            df.loc[i, 'Cas_Vykonu_Do'] = ''
            df.loc[i, 'Prestavka_Trvanie'] = '00:00:00'
            df.loc[i, 'Popis_Cinnosti'] = 'DOVOLENKA'
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '08:00:00'
            df.loc[i, 'Miesto_Vykonu'] = ''
            df.loc[i, 'PH_Projekt_POO'] = '00:00:00'
            df.loc[i, 'PH_Riesenie_POO'] = '00:00:00'
            df.loc[i, 'PH_Mimo_Projekt_POO'] = '00:00:00'
            df.loc[i, 'SPOLU'] = '08:00:00'
            logging.info(f"Applied full-day vacation to day {day_num}")
        elif vac_type == 'morning':
            # Morning is vacation → work afternoon
            df.loc[i, 'Cas_Vykonu_Od'] = '13:00:00'
            df.loc[i, 'Cas_Vykonu_Do'] = '17:00:00'
            df.loc[i, 'Prestavka_Trvanie'] = '00:00:00'
            df.loc[i, 'Popis_Cinnosti'] = activity_text
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '04:00:00'
            df.loc[i, 'Miesto_Vykonu'] = work_location
            df.loc[i, 'PH_Projekt_POO'] = '00:00:00'
            df.loc[i, 'PH_Riesenie_POO'] = '00:00:00'
            df.loc[i, 'PH_Mimo_Projekt_POO'] = '00:00:00'
            df.loc[i, 'SPOLU'] = '04:00:00'
            logging.info(f"Applied morning-vacation (work afternoon) to day {day_num}")
        elif vac_type == 'afternoon':
            # Afternoon is vacation → work morning
            df.loc[i, 'Cas_Vykonu_Od'] = '09:00:00'
            df.loc[i, 'Cas_Vykonu_Do'] = '13:00:00'
            df.loc[i, 'Prestavka_Trvanie'] = '00:00:00'
            df.loc[i, 'Popis_Cinnosti'] = activity_text
            df.loc[i, 'Pocet_Odpracovanych_Hodin'] = '04:00:00'
            df.loc[i, 'Miesto_Vykonu'] = work_location
            df.loc[i, 'PH_Projekt_POO'] = '00:00:00'
            df.loc[i, 'PH_Riesenie_POO'] = '00:00:00'
            df.loc[i, 'PH_Mimo_Projekt_POO'] = '00:00:00'
            df.loc[i, 'SPOLU'] = '04:00:00'
            logging.info(f"Applied afternoon-vacation (work morning) to day {day_num}")

    return df


def load_vacations(json_path: str) -> dict:
    """Load vacation data from JSON file.

    Returns:
        Dict with keys: month, year, vacations (employee_name -> list of vacation entries)
    """
    import json
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def match_vacation_to_sheet(sheet_name: str, vacations: dict) -> Optional[List]:
    """Match a target sheet name to vacation entries using normalized name comparison.

    Returns the vacation day list if matched, None otherwise.
    """
    from src.sheet_mapper import _normalize_name
    norm_sheet = _normalize_name(sheet_name)
    for emp_name, days in vacations.items():
        if _normalize_name(emp_name) == norm_sheet:
            return days
    return None


def update_daily_rows(ws, df_target: pd.DataFrame, data_start_row: int):
    """Update daily rows in the target worksheet."""
    col_mappings = {
        'Datum': 1,
        'Cas_Vykonu_Od': 2,
        'Cas_Vykonu_Do': 3,
        'Prestavka_Trvanie': 4,
        'Popis_Cinnosti': 5,
        'Pocet_Odpracovanych_Hodin': 9,
        'Miesto_Vykonu': 10,
        'PH_Projekt_POO': 11,
        'PH_Riesenie_POO': 12,
        'PH_Mimo_Projekt_POO': 13,
        'SPOLU': 14
    }
    
    # Store original merged ranges to restore later
    original_ranges = list(ws.merged_cells.ranges)
    unmerged_coords = set()
    
    try:
        for i in range(31):
            target_row = data_start_row + i
            
            # Unmerge cells for this row if needed
            for merged_range in original_ranges:
                if merged_range.min_row <= target_row <= merged_range.max_row:
                    coord = merged_range.coord
                    if coord not in unmerged_coords:
                        ws.unmerge_cells(coord)
                        unmerged_coords.add(coord)
                        logging.debug(f"Unmerging {coord} for row {target_row}")
            
            # Update cells
            for col_name, col_num in col_mappings.items():
                val = df_target.iloc[i][col_name]
                # sanitize time-like fields before writing to Excel
                if col_name in ('Pocet_Odpracovanych_Hodin', 'Prestavka_Trvanie', 'PH_Projekt_POO', 'PH_Riesenie_POO', 'PH_Mimo_Projekt_POO', 'SPOLU'):
                    # reuse the sanitizer defined in outer scope by calling a small inline function
                    def _sanitize_for_write(v):
                        if v is None or (isinstance(v, float) and pd.isna(v)):
                            return ''
                        if isinstance(v, set) and len(v) == 1:
                            v = next(iter(v))
                        if isinstance(v, str) and v.strip() == '-':
                            return ''
                        return v
                    val = _sanitize_for_write(val)
                if pd.isna(val) or val == '-':
                    val = ''
                ws.cell(row=target_row, column=col_num, value=val)
                
                # Clear merged cells for description if it has content
                if col_name == 'Popis_Cinnosti' and val != '':
                    for c in [6, 7, 8]:
                        ws.cell(row=target_row, column=c, value='')
        
        # Re-merge cells that were unmerged
        for coord in unmerged_coords:
            ws.merge_cells(coord)
        logging.info(f"Re-merged {len(unmerged_coords)} ranges")
        
    except Exception as e:
        logging.error(f"Error during row update: {e}")


def recalculate_summary(df_target: pd.DataFrame, ws):
    """Recalculate and update summary row."""
    # Count work days
    try:
        work_days = len(df_target[df_target['SPOLU'] != '00:00:00'])
    except Exception as e:
        logging.warning(f"Error counting work days: {e}")
        work_days = 0
    
    # Sum total hours
    total_td = timedelta()
    for value in df_target['SPOLU']:
        if value not in (None, '00:00:00', ''):
            try:
                h, m, s = map(int, str(value).split(':'))
                total_td += timedelta(hours=h, minutes=m, seconds=s)
            except (ValueError, AttributeError, TypeError) as e:
                logging.warning(f"Could not parse SPOLU value '{value}': {e}")
    
    # Format total time
    try:
        total_seconds = total_td.total_seconds()
        hours = int(abs(total_seconds) // 3600)
        minutes = int((abs(total_seconds) % 3600) // 60)
        seconds = int(abs(total_seconds) % 60)
        total_time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    except Exception as e:
        logging.warning(f"Error formatting total time: {e}")
        total_time_str = '00:00:00'
    
    # Update summary cell (typically row 57, column 14)
    try:
        ws.cell(row=57, column=14, value=total_time_str)
        logging.info(f"Summary updated: {total_time_str} in N57")
    except Exception as e:
        logging.error(f"Error updating summary cell: {e}")
    
    return f"{work_days} days, {total_time_str}", total_time_str


def save_and_validate(wb, df_target: Optional[pd.DataFrame], backup_path: str, output_dir: str, dry_run: bool):
    """Save workbook and generate output files."""
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate timestamp for output files
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"updated_{timestamp}.xlsx")
    
    if dry_run:
        logging.info("Dry run: skipping workbook save")
        wb.close()
        return

    try:
        wb.save(output_path)
        logging.info(f"Workbook saved to {output_path}")

        # Save CSV for audit only if df_target is provided
        if df_target is not None:
            # Ensure transformed CSV subdirectory exists and write CSVs there
            transformed_dir = os.path.join(output_dir, 'transformed')
            os.makedirs(transformed_dir, exist_ok=True)
            csv_path = os.path.join(transformed_dir, f"transformed_data_{timestamp}.csv")
            df_target.to_csv(csv_path, index=False)
            logging.info(f"Transformed CSV saved to {csv_path}")

        wb.close()
        logging.info("Workbook closed successfully")

    except PermissionError as e:
        logging.error(f"Permission error saving workbook: {e}. Please close Excel file and retry.")
        try:
            wb.close()
        except Exception:
            pass
    except Exception as e:
        logging.error(f"Error saving workbook: {e}")
        if backup_path and os.path.exists(backup_path):
            logging.info(f"Backup available at: {backup_path}")
        try:
            wb.close()
        except Exception:
            pass
        raise


def main():
    """Main function to orchestrate the update process."""
    args = parse_args()

    SLOVAK_MONTHS = {
        'január': 1, 'február': 2, 'marec': 3, 'apríl': 4,
        'máj': 5, 'jún': 6, 'júl': 7, 'august': 8,
        'september': 9, 'október': 10, 'november': 11, 'december': 12
    }

    # Dispatch: vacations-only mode vs source-based mode
    if args.vacations_json and not args.source_excel:
        return _main_vacations_mode(args, SLOVAK_MONTHS)
    elif args.source_excel:
        return _main_source_mode(args, SLOVAK_MONTHS)
    else:
        raise ValueError("Either --source-excel or --vacations-json must be provided")


def _main_vacations_mode(args, SLOVAK_MONTHS):
    """Process all target sheets using vacation JSON (no source attendance file)."""
    if not args.month or not args.year:
        raise ValueError("--month and --year are required when using --vacations-json")

    month_num = SLOVAK_MONTHS.get(args.month.lower())
    if not month_num:
        raise ValueError(f"Unknown month: '{args.month}'")

    logging.info("Starting vykaz update (vacations-only mode)")

    # Load vacation data
    vac_data = load_vacations(args.vacations_json)
    vacations = vac_data.get("vacations", {})
    logging.info(f"Loaded vacations for {len(vacations)} employees")

    # Load config for protected sheets
    mappings_config = sheet_mapper.load_mappings_config()
    protected_sheets = set(mappings_config.get("protected_sheets", []))

    target_file_to_process = args.target_excel

    # Create backup
    backup_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(os.path.dirname(target_file_to_process), 'backup')
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"backup_{backup_timestamp}.xlsx")

    if not args.dry_run and os.path.exists(target_file_to_process):
        shutil.copy(target_file_to_process, backup_path)
        logging.info(f"Backup created: {backup_path}")
    else:
        backup_path = None

    # Load target workbook
    wb = load_workbook(target_file_to_process)
    from src.extractor_utils import STRATEGY_REGISTRY
    target_strategy = STRATEGY_REGISTRY["target"]
    data_start_row = target_strategy["start_row_strategy"](None)

    processed_sheets = 0
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() in protected_sheets:
            logging.info(f"Skipping protected sheet: {sheet_name}")
            continue

        # Match vacation days for this employee
        vac_days = match_vacation_to_sheet(sheet_name, vacations)
        if vac_days:
            logging.info(f"Processing {sheet_name} with {len(vac_days)} vacation entries")
        else:
            logging.info(f"Processing {sheet_name} (no vacations)")

        df = generate_data_with_vacations(
            args.year, month_num, args.activity_text, args.work_location,
            vacation_days=vac_days
        )

        ws = wb[sheet_name]

        # Update month
        if args.month:
            try:
                ws['E13'] = args.month
            except Exception as e:
                logging.warning(f"Could not update month in E13 for {sheet_name}: {e}")

        update_daily_rows(ws, df, data_start_row)
        summary_text, _ = recalculate_summary(df, ws)
        logging.info(f"Summary for {sheet_name}: {summary_text}")

        # Save transformed CSV
        try:
            transformed_dir = os.path.join(args.output_dir, 'transformed')
            os.makedirs(transformed_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = sheet_name.replace(' ', '_').replace('/', '_')
            csv_path = os.path.join(transformed_dir, f"transformed_{safe_name}_{ts}.csv")
            df.to_csv(csv_path, index=False)
            logging.info(f"Transformed CSV saved to {csv_path}")
        except Exception as e:
            logging.warning(f"Could not save CSV for {sheet_name}: {e}")

        processed_sheets += 1

    logging.info(f"Total sheets processed: {processed_sheets}")
    save_and_validate(wb, None, backup_path, args.output_dir, args.dry_run)
    logging.info("Process completed successfully")


def _main_source_mode(args, SLOVAK_MONTHS):
    """Original source-based processing mode."""
    try:
        logging.info("Starting vykaz update process")

        # Step 1: Create sheet mappings between source and target
        logging.info("Creating sheet mappings...")
        source_sheets = sheet_mapper.extract_sheet_names(args.source_excel)
        source_sheets = sheet_mapper.filter_instruction_sheets(source_sheets)
        target_sheets = sheet_mapper.extract_sheet_names(args.target_excel)

        if not source_sheets:
            raise ValueError(f"Could not extract sheet names from source file: {args.source_excel}")
        if not target_sheets:
            raise ValueError(f"Could not extract sheet names from target file: {args.target_excel}")

        mapping, unmatched_source, unmatched_target = sheet_mapper.create_mapping(source_sheets, target_sheets)
        logging.info(f"Created mappings for {len(mapping)} sheets")

        # Load config for protected sheets
        mappings_config = sheet_mapper.load_mappings_config()
        protected_sheets = mappings_config.get("protected_sheets", [])

        if unmatched_source:
            logging.warning(f"Unmatched source sheets: {unmatched_source}")
        if unmatched_target:
            logging.warning(f"Unmatched target sheets: {unmatched_target}")

        # Separate unmatched targets into contractors (to fill) and protected (to keep as-is)
        contractor_names, protected_names = sheet_mapper.filter_protected_from_unmatched(
            unmatched_target, protected_sheets
        )
        if contractor_names:
            logging.info(f"Contractor sheets (will fill with 8h/day): {contractor_names}")
        if protected_names:
            logging.info(f"Protected sheets (kept as-is): {protected_names}")

        # Step 1.5: No sheets are removed — contractors and protected are kept
        cleaned_target_path = args.target_excel

        # Step 1.6: Sort target sheets based on source sheet order
        target_file_to_process = cleaned_target_path
        if not args.no_sort_target:
            logging.info("Sorting target sheets based on source sheet order...")
            sorted_target_path = sheet_mapper.sort_target_sheets_by_source_order(
                args.source_excel,
                cleaned_target_path,
                mapping,
                save_sorted=True
            )
            if sorted_target_path:
                logging.info(f"Sorted target workbook saved to: {sorted_target_path}")
                target_file_to_process = sorted_target_path
        else:
            logging.info("Skipping target sorting (disabled by --no-sort-target)")

        # Step 2: Create backup of target file
        backup_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.join(os.path.dirname(target_file_to_process), 'backup')
        os.makedirs(backup_dir, exist_ok=True)
        backup_path = os.path.join(backup_dir, f"backup_{backup_timestamp}.xlsx")

        if not args.dry_run and os.path.exists(target_file_to_process):
            shutil.copy(target_file_to_process, backup_path)
            logging.info(f"Backup created: {backup_path}")
        else:
            backup_path = None
            if args.dry_run:
                logging.info("Dry run: skipping backup creation")

        # Step 3: Load target workbook once
        logging.info("Loading target Excel...")
        wb = load_workbook(target_file_to_process)

        # Step 4: Process each mapped sheet
        processed_sheets = 0
        for source_sheet, target_sheet in mapping.items():
            if target_sheet == '-':
                logging.info(f"Skipping unmapped source sheet: {source_sheet}")
                continue

            logging.info(f"Processing sheet mapping: {source_sheet} -> {target_sheet}")

            try:
                # Extract source data for this specific sheet
                logging.info(f"Extracting source data from sheet: {source_sheet}")
                df_source = extract_source_data(args.source_excel, source_sheet)
                logging.info(f"Extracted {len(df_source)} rows from {source_sheet}")

                # Transform data
                logging.info(f"Transforming data for sheet: {target_sheet}")
                df_target = source_to_target(
                    df_source,
                    args.activity_text,
                    args.work_location
                )
                logging.info("Data transformation completed")

                # Get target worksheet
                if target_sheet not in wb.sheetnames:
                    logging.error(f"Target sheet '{target_sheet}' not found in workbook")
                    continue

                ws = wb[target_sheet]

                # Find data start row using the target strategy
                from src.extractor_utils import STRATEGY_REGISTRY
                target_strategy = STRATEGY_REGISTRY["target"]
                data_start_row = target_strategy["start_row_strategy"](None)
                logging.info(f"Using data start row: {data_start_row}")

                # Update month if provided
                if args.month:
                    try:
                        ws['E13'] = args.month
                        logging.info(f"Updated cell E13 with month: {args.month}")
                    except Exception as e:
                        logging.warning(f"Could not update month in E13: {e}")

                # Update daily rows
                logging.info(f"Updating daily rows in sheet: {target_sheet}")
                update_daily_rows(ws, df_target, data_start_row)

                # Recalculate summary
                logging.info(f"Recalculating summary for sheet: {target_sheet}")
                summary_text, total_time = recalculate_summary(df_target, ws)
                logging.info(f"Summary for {target_sheet}: {summary_text}")
                # Save transformed CSV for this sheet into transformed subfolder
                try:
                    transformed_dir = os.path.join(args.output_dir, 'transformed')
                    os.makedirs(transformed_dir, exist_ok=True)
                    def _normalize_df_times(df):
                        def _fix(v):
                            if v is None or (isinstance(v, float) and pd.isna(v)):
                                return ''
                            if isinstance(v, set) and len(v) == 1:
                                return next(iter(v))
                            if isinstance(v, str) and v.strip() == '-':
                                return ''
                            return v
                        for c in ['Pocet_Odpracovanych_Hodin','Prestavka_Trvanie','PH_Projekt_POO','PH_Riesenie_POO','PH_Mimo_Projekt_POO','SPOLU']:
                            if c in df.columns:
                                df[c] = df[c].apply(_fix).astype(str)
                        return df

                    csv_df = _normalize_df_times(df_target.copy())
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    safe_name = target_sheet.replace(' ', '_').replace('/', '_')
                    csv_path = os.path.join(transformed_dir, f"transformed_{safe_name}_{ts}.csv")
                    csv_df.to_csv(csv_path, index=False)
                    logging.info(f"Transformed CSV saved to {csv_path}")
                except Exception as e:
                    logging.warning(f"Could not save transformed CSV for {target_sheet}: {e}")

                processed_sheets += 1

            except Exception as e:
                logging.error(f"Error processing sheet {source_sheet} -> {target_sheet}: {e}")
                continue

        logging.info(f"Successfully processed {processed_sheets} employee sheets")

        # Step 4b: Process contractor sheets (unmatched targets that aren't protected)
        if contractor_names and args.month and args.year:
            month_num = SLOVAK_MONTHS.get(args.month.lower())
            if not month_num:
                logging.warning(f"Could not resolve month '{args.month}' — skipping contractors")
            else:
                logging.info(f"Processing {len(contractor_names)} contractor sheets...")
                from src.extractor_utils import STRATEGY_REGISTRY
                target_strategy = STRATEGY_REGISTRY["target"]

                for contractor_sheet in contractor_names:
                    if contractor_sheet not in wb.sheetnames:
                        logging.warning(f"Contractor sheet '{contractor_sheet}' not found in workbook, skipping")
                        continue

                    logging.info(f"Processing contractor: {contractor_sheet}")
                    df_contractor = generate_contractor_data(
                        args.year, month_num, args.activity_text, args.work_location
                    )

                    ws = wb[contractor_sheet]
                    data_start_row = target_strategy["start_row_strategy"](None)

                    if args.month:
                        try:
                            ws['E13'] = args.month
                        except Exception as e:
                            logging.warning(f"Could not update month in E13: {e}")

                    update_daily_rows(ws, df_contractor, data_start_row)
                    summary_text, _ = recalculate_summary(df_contractor, ws)
                    logging.info(f"Summary for contractor {contractor_sheet}: {summary_text}")

                    try:
                        transformed_dir = os.path.join(args.output_dir, 'transformed')
                        os.makedirs(transformed_dir, exist_ok=True)
                        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                        safe_name = contractor_sheet.replace(' ', '_').replace('/', '_')
                        csv_path = os.path.join(transformed_dir, f"transformed_{safe_name}_{ts}.csv")
                        df_contractor.to_csv(csv_path, index=False)
                        logging.info(f"Transformed CSV saved to {csv_path}")
                    except Exception as e:
                        logging.warning(f"Could not save CSV for {contractor_sheet}: {e}")

                    processed_sheets += 1
        elif contractor_names:
            logging.warning("Skipping contractors: --month and --year are required")

        logging.info(f"Total sheets processed: {processed_sheets}")

        # Step 5: Save and validate
        logging.info("Saving workbook...")
        save_and_validate(wb, None, backup_path, args.output_dir, args.dry_run)
        logging.info("Process completed successfully")

    except Exception as e:
        logging.error(f"An error occurred: {e}")
        raise


if __name__ == "__main__":
    main()

