import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_cell_by_text(sheet, search_text):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_text:
                return cell.row, cell.column, get_column_letter(cell.column)

def process_excel(file_path, source_target):
    wb = load_workbook(file_path)
    sheet = wb.active  # Assuming first sheet

    if source_target == "source":
        datum_text = "Dátum"
        hours_text = "odpracovaný čas"
    elif source_target == "target":
        datum_text = "Dátum"
        hours_text = "Počet odpracovaných hodín*"

    datum_pos = find_cell_by_text(sheet, datum_text)
    hours_pos = find_cell_by_text(sheet, hours_text)

    return datum_pos, hours_pos

# Source file
source_datum, source_hours = process_excel("ronec_dochadzka.xlsx", "source")

# Target file
target_datum, target_hours = process_excel("ronec_vykaz.xlsx", "target")

print("Source File (ronec_dochadzka.xlsx):")
print("\"Dátum\" found at:")
if source_datum:
    print(f"  Row: {source_datum[0]}, Column: {source_datum[2]} ({source_datum[1]})")
else:
    print("  Not found")

print("\"odpracovaný čas\" found at:")
if source_hours:
    print(f"  Row: {source_hours[0]}, Column: {source_hours[2]} ({source_hours[1]})")
else:
    print("  Not found")

print("\nTarget File (ronec_vykaz.xlsx):")
print("\"Dátum\" found at:")
if target_datum:
    print(f"  Row: {target_datum[0]}, Column: {target_datum[2]} ({target_datum[1]})")
else:
    print("  Not found")

print("\"Počet odpracovaných hodín*\" found at:")
if target_hours:
    print(f"  Row: {target_hours[0]}, Column: {target_hours[2]} ({target_hours[1]})")
else:
    print("  Not found")