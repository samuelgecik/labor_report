import os
import json
import tempfile
import subprocess
from pathlib import Path
from openpyxl import Workbook, load_workbook

# New tests aligned with refactored runtime mapping pipeline (Step 15)
# Focus: sheet creation, 31 rows, summary update, cleaned target usage.

DAILY_START_ROW = 26
DAILY_ROW_COUNT = 31
SUMMARY_ROW = 57


def _make_source_wb(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ing. Test Person'
    # Provide minimal data for first few days; transform logic pads to 31
    for day in range(1, 5):
        ws.cell(row=day, column=2, value='09:00')  # start
        ws.cell(row=day, column=3, value='17:00')  # end
        ws.cell(row=day, column=4, value='60')     # break minutes
        ws.cell(row=day, column=6, value='08:00')  # worked hours
    wb.save(path)
    wb.close()


def _make_target_wb(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = 'TemplateSheet'
    wb.save(path)
    wb.close()


def _run_script(args: list[str]):
    cmd = ['python3', '-m', 'src.update_vykaz'] + args
    return subprocess.run(cmd, capture_output=True, text=True)


def test_integration_creates_sheet_and_rows():
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, 'source.xlsx')
        target = os.path.join(tmp, 'target.xlsx')
        _make_source_wb(source)
        _make_target_wb(target)
        res = _run_script(['--source-excel', source, '--target-excel', target])
        assert res.returncode == 0, res.stderr
        wb = load_workbook(target)
        assert 'Test Person' in wb.sheetnames  # mapped name
        ws = wb['Test Person']
        for i in range(DAILY_ROW_COUNT):
            assert ws.cell(row=DAILY_START_ROW + i, column=1).value == f"{i+1}."
        wb.close()


def test_summary_row_updated():
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, 'source.xlsx')
        target = os.path.join(tmp, 'target.xlsx')
        _make_source_wb(source)
        _make_target_wb(target)
        res = _run_script(['--source-excel', source, '--target-excel', target])
        assert res.returncode == 0, res.stderr
        wb = load_workbook(target, data_only=True)
        ws = wb['Test Person']
        assert ws.cell(row=SUMMARY_ROW, column=14).value is not None
        wb.close()


def test_clean_target_path_used():
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, 'source.xlsx')
        target = os.path.join(tmp, 'target.xlsx')
        _make_source_wb(source)
        _make_target_wb(target)
        wb = load_workbook(target)
        wb.create_sheet('ExtraUnmatchedSheet')
        wb.save(target)
        wb.close()
        res = _run_script(['--source-excel', source, '--target-excel', target, '--clean-target'])
        assert res.returncode == 0, res.stderr
        cleaned_path = Path(target).with_name(Path(target).stem + '_cleaned.xlsx')
        assert cleaned_path.exists()
        wb2 = load_workbook(cleaned_path)
        assert 'ExtraUnmatchedSheet' not in wb2.sheetnames
        wb2.close()


def test_activities_override():
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, 'source.xlsx')
        target = os.path.join(tmp, 'target.xlsx')
        activities_json = os.path.join(tmp, 'activities.json')
        _make_source_wb(source)
        _make_target_wb(target)
        override_text = 'Custom Activity Text X'
        with open(activities_json, 'w', encoding='utf-8') as f:
            json.dump({'Ing. Test Person': override_text}, f)
        res = _run_script(['--source-excel', source, '--target-excel', target, '--activities-json', activities_json])
        assert res.returncode == 0, res.stderr
        wb = load_workbook(target)
        ws = wb['Test Person']
        assert ws.cell(row=DAILY_START_ROW, column=5).value == override_text
        wb.close()


def test_dry_run_no_write():
    with tempfile.TemporaryDirectory() as tmp:
        source = os.path.join(tmp, 'source.xlsx')
        target = os.path.join(tmp, 'target.xlsx')
        _make_source_wb(source)
        _make_target_wb(target)
        before = os.path.getmtime(target)
        res = _run_script(['--source-excel', source, '--target-excel', target, '--dry-run'])
        assert res.returncode == 0, res.stderr
        after = os.path.getmtime(target)
        assert before == after, 'File should not be modified in dry-run'

if __name__ == '__main__':
    import pytest
    raise SystemExit(pytest.main([__file__, '-q']))