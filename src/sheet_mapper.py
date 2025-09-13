import openpyxl
import difflib
import unicodedata
import argparse
import json
import os
from datetime import date

source_path = '/home/gobi/vykazy/data/input/source_test.xlsx'
target_path = '/home/gobi/vykazy/data/output/updated_20250913_193959.xlsx'
# Central list of instruction sheet names to exclude in mappings
INSTRUCTION_SHEET_NAMES = {"Inštrukcie k vyplneniu PV", "Instrukcie k vyplneniu PV"}

def extract_sheet_names(path):
    try:
        wb = openpyxl.load_workbook(path)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except FileNotFoundError:
        print(f"File not found: {path}")
        return []
    except Exception as e:
        print(f"Error loading {path}: {e}")
        return []

def filter_instruction_sheets(sheet_names):
    """Return a new list without instruction sheets.

    Uses INSTRUCTION_SHEET_NAMES to filter out non-data sheets.
    """
    return [s for s in sheet_names if s not in INSTRUCTION_SHEET_NAMES]

def _remove_titles(name):
    prefixes = ['Ing.', 'Bc.', 'Mgr.', 'PhD.', 'prof.', 'MUDr.', 'RNDr.']
    for prefix in prefixes:
        if name.startswith(prefix + ' '):
            name = name[len(prefix) + 1:]
            break
    return name

def _normalize_name(name):
    name = _remove_titles(name)
    name = unicodedata.normalize('NFD', name).encode('ascii', 'ignore').decode('ascii')
    return name.strip().lower()

def create_mapping(source_sheets, target_sheets):
    mapping = {}
    unmatched_source = []
    norm_targets = [_normalize_name(t) for t in target_sheets]
    used_targets = set()
    for source in source_sheets:
        norm_source = _normalize_name(source)
        if norm_source in norm_targets:
            matched = target_sheets[norm_targets.index(norm_source)]
            used_targets.add(matched)
        else:
            close = difflib.get_close_matches(norm_source, norm_targets, n=1, cutoff=0.8)
            if close:
                matched = target_sheets[norm_targets.index(close[0])]
                used_targets.add(matched)
            else:
                matched = '-'
                unmatched_source.append(f"{source} -> -")
        mapping[source] = matched
    unmatched_target = [t for t in target_sheets if t not in used_targets and t != '-']
    unmatched_target = [f"{t} -> -" for t in unmatched_target]
    return mapping, unmatched_source, unmatched_target

def remove_unmatched_target_sheets(target_path, unmatched_target):
    """Create a cleaned copy of target workbook without unmatched sheets.

    Ensures at least one sheet remains: if all target sheets are marked unmatched,
    the first one is kept as a template and the rest are removed.

    Returns path to the cleaned workbook ("*_cleaned.xlsx").
    """
    wb = openpyxl.load_workbook(target_path)
    sheets_to_remove = [name.split(" -> -")[0] for name in unmatched_target]
    # If all sheets would be removed, keep the first as template
    if sheets_to_remove and set(sheets_to_remove) >= set(wb.sheetnames):
        keep = sheets_to_remove[0]
        for sheet_name in sheets_to_remove[1:]:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
    else:
        for sheet_name in sheets_to_remove:
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
    base, ext = os.path.splitext(target_path)
    cleaned_path = base + '_cleaned' + ext
    wb.save(cleaned_path)
    wb.close()
    return cleaned_path

def sort_target_sheets_by_source_order(source_path, target_path, mapping=None, save_sorted=True):
    """Sort sheets in target workbook based on the order of sheets in source workbook.
    
    Args:
        source_path: Path to source Excel file
        target_path: Path to target Excel file
        mapping: Optional mapping dict from create_mapping() function. If not provided,
                it will be generated automatically.
        save_sorted: If True, saves the sorted workbook with "_sorted" suffix
        
    Returns:
        str: Path to the sorted workbook file if save_sorted=True, else None
    """
    try:
        # Load both workbooks
        source_wb = openpyxl.load_workbook(source_path)
        target_wb = openpyxl.load_workbook(target_path)
        
        # Get sheet names from source (filtered)
        source_sheets = filter_instruction_sheets(source_wb.sheetnames)
        target_sheets = target_wb.sheetnames
        
        # Create mapping if not provided
        if mapping is None:
            mapping, _, _ = create_mapping(source_sheets, target_sheets)
        
        # Create ordered list of target sheets based on source order
        ordered_target_sheets = []
        unordered_sheets = list(target_sheets)  # Copy to track remaining sheets
        
        # First, add sheets in source order (if they exist in target)
        for source_sheet in source_sheets:
            target_sheet = mapping.get(source_sheet, '-')
            if target_sheet != '-' and target_sheet in unordered_sheets:
                ordered_target_sheets.append(target_sheet)
                unordered_sheets.remove(target_sheet)
        
        # Add any remaining target sheets that weren't mapped
        ordered_target_sheets.extend(unordered_sheets)
        
        # Reorder sheets in target workbook
        # OpenPyxl doesn't have direct sheet reordering, so we need to move sheets
        for i, sheet_name in enumerate(ordered_target_sheets):
            if sheet_name in target_wb.sheetnames:
                sheet = target_wb[sheet_name]
                # Move sheet to the correct position
                target_wb.move_sheet(sheet, offset=i - target_wb.index(sheet))
        
        source_wb.close()
        
        if save_sorted:
            # Save sorted workbook
            base, ext = os.path.splitext(target_path)
            sorted_path = base + '_sorted' + ext
            target_wb.save(sorted_path)
            target_wb.close()
            print(f"Sorted target workbook saved to: {sorted_path}")
            return sorted_path
        else:
            target_wb.close()
            return None
            
    except FileNotFoundError as e:
        print(f"File not found: {e}")
        return None
    except Exception as e:
        print(f"Error sorting sheets: {e}")
        return None

def save_mapping_json(mapping, unmatched_source, unmatched_target, output_dir, user_path, activities=None, metadata=None):
    """Save runtime mapping data to JSON file with activities and metadata.

    Args:
        mapping: Dict mapping source sheet names to target sheet names
        unmatched_source: List of unmatched source sheet entries  
        unmatched_target: List of unmatched target sheet entries
        output_dir: Output directory for JSON file
        user_path: User-specified path or True/False for auto-naming
        activities: Optional dict of activity overrides per sheet
        metadata: Optional dict of metadata per sheet

    Returns path to saved JSON file.
    """
    from datetime import datetime
    
    os.makedirs(output_dir, exist_ok=True)
    if isinstance(user_path, str) and user_path not in ("True", "true", "FALSE", "False"):
        out_path = user_path
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(output_dir, f"mappings_runtime_{timestamp}.json")
    
    payload = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "mapping": mapping,
        "unmatched_source": unmatched_source,
        "unmatched_target": unmatched_target,
    }
    if activities:
        payload["activities"] = activities
    if metadata:
        payload["metadata"] = metadata
    
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    
    print(f"Runtime mapping JSON saved: {out_path}")
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Map sheet names from source to target Excel files.')
    parser.add_argument('--source', default=source_path, help='Path to source Excel file')
    parser.add_argument('--target', default=target_path, help='Path to target Excel file')
    parser.add_argument('--clean-target', action='store_true', default=True, help='Remove unmatched target sheets and save cleaned file')
    parser.add_argument('--sort-target', action='store_true', default=False, help='Sort target sheets based on source sheet order')
    args = parser.parse_args()

    source_sheets = extract_sheet_names(args.source)
    source_sheets = filter_instruction_sheets(source_sheets)
    target_sheets = extract_sheet_names(args.target)

    if not source_sheets:
        print(f"Could not extract sheet names from source file: {args.source}")
        exit(1)
    if not target_sheets:
        print(f"Could not extract sheet names from target file: {args.target}")
        exit(1)

    mapping, unmatched_source, unmatched_target = create_mapping(source_sheets, target_sheets)

    print("Sheet name mappings:")
    for source, target in mapping.items():
        print(f"{source} -> {target}")
    if unmatched_source:
        print("Unmatched source sheets:")
        for unmatched in unmatched_source:
            print(unmatched)
    if unmatched_target:
        print("Unmatched target sheets:")
        for unmatched in unmatched_target:
            print(unmatched)
    
    # Handle sheet sorting if requested
    if args.sort_target:
        sorted_file = sort_target_sheets_by_source_order(args.source, args.target, mapping)
        if sorted_file:
            print(f"Target sheets sorted and saved to: {sorted_file}")
    
    if args.clean_target and unmatched_target:
        cleaned_file = remove_unmatched_target_sheets(args.target, unmatched_target)
        print(f"Cleaned target file saved to {cleaned_file}")
    
    data = {
        "mappings": mapping,
        "unmatched_source": unmatched_source,
        "unmatched_target": unmatched_target
    }
    filename = f'data/mappings_{date.today().isoformat()}.json'
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"Mappings saved to {filename}")