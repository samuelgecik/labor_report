#!/usr/bin/env python3
"""Append randomized task descriptions to labor report sheets based on working position.

This script augments each sheet in a workbook (e.g. ``data/input/jun_2025.xlsx``)
by appending concrete tasks sourced from ``data/ulohy.json`` to the activity
column (column ``E``). Tasks are selected according to the working position
found in cell ``E12`` and are assigned in multi-day blocks to simulate work
spanning several days or weeks.
"""
from __future__ import annotations

import argparse
import json
import logging
import random
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


DEFAULT_MIN_BLOCK = 3
DEFAULT_MAX_BLOCK = 10
DEFAULT_MONTH_NAME = "júl"
DEFAULT_YEAR = 2025
COLUMN_ACTIVITY = 5  # Column E
COLUMN_DAY = 1       # Column A
FIRST_DATA_ROW = 26
DAY_COUNT = 31


@dataclass(frozen=True)
class DayEntry:
    """Represents a row that holds an activity description for a specific day."""

    row_index: int
    day_number: int
    base_text: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append randomized tasks into labor report workbook activities"
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path("data/input/jun_2025.xlsx"),
        help="Input workbook to update"
    )
    parser.add_argument(
        "--tasks-json",
        type=Path,
        default=Path("data/ulohy.json"),
        help="Path to JSON containing tasks grouped by working position"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/output/jun_2025_with_tasks.xlsx"),
        help="Output path for the augmented workbook"
    )
    parser.add_argument(
        "--min-block",
        type=int,
        default=DEFAULT_MIN_BLOCK,
        help="Minimum consecutive-day block to assign the same tasks"
    )
    parser.add_argument(
        "--max-block",
        type=int,
        default=DEFAULT_MAX_BLOCK,
        help="Maximum consecutive-day block to assign the same tasks"
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Optional random seed for reproducibility"
    )
    parser.add_argument(
        "--month-name",
        type=str,
        default=DEFAULT_MONTH_NAME,
        help="Month name to use in generated text (default: jún)"
    )
    parser.add_argument(
        "--year",
        type=int,
        default=DEFAULT_YEAR,
        help="Year to use in generated text (default: 2025)"
    )
    parser.add_argument(
        "--skip-sheets",
        type=str,
        nargs="*",
        default=("Inštrukcie",),
        help="Sheet name prefixes to skip (case insensitive)"
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite the output file if it already exists"
    )
    return parser.parse_args()


def load_tasks(tasks_json: Path) -> Dict[str, List[str]]:
    with tasks_json.open("r", encoding="utf-8") as fh:
        raw = json.load(fh)

    task_pool: Dict[str, set[str]] = defaultdict(set)

    for balík in raw:
        for position in balík.get("pracovne_pozicie", []):
            pozicia = position.get("pozicia")
            if not pozicia:
                continue
            for task in position.get("ulohy", []):
                cleaned = task.strip()
                if cleaned:
                    task_pool[pozicia.strip()].add(cleaned)

    if not task_pool:
        raise ValueError(f"No tasks found in {tasks_json}")

    return {key: sorted(values) for key, values in task_pool.items()}


def extract_day_entries(ws: Worksheet) -> List[DayEntry]:
    entries: List[DayEntry] = []
    for offset in range(DAY_COUNT):
        row_idx = FIRST_DATA_ROW + offset
        activity_cell = ws.cell(row=row_idx, column=COLUMN_ACTIVITY)
        raw_value = activity_cell.value
        if not isinstance(raw_value, str) or not raw_value.strip():
            continue
        # skip vacation rows
        if raw_value.strip().upper() == "DOVOLENKA":
            continue
        # avoid re-appending if script already processed this cell
        if " - Úlohy:" in raw_value:
            logger.debug("Skipping row %s in sheet %s because it's already augmented", row_idx, ws.title)
            continue

        day_cell_value = ws.cell(row=row_idx, column=COLUMN_DAY).value
        day_number = parse_day_number(day_cell_value)
        if day_number is None:
            logger.debug("Skipping row %s in sheet %s due to missing day number", row_idx, ws.title)
            continue

        entries.append(
            DayEntry(row_index=row_idx, day_number=day_number, base_text=raw_value.rstrip())
        )
    return entries


def parse_day_number(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        text = str(value).strip()
        if not text:
            return None
        # Expect format like "1." or "01."
        if text.endswith("."):
            text = text[:-1]
        return int(text)
    except (TypeError, ValueError):
        return None


def partition_entries(entries: Sequence[DayEntry], min_len: int, max_len: int) -> List[List[DayEntry]]:
    if not entries:
        return []
    if min_len < 1 or max_len < min_len:
        raise ValueError("Invalid block length configuration")

    blocks: List[List[DayEntry]] = []
    i = 0
    count = len(entries)
    while i < count:
        remaining = count - i
        if remaining <= min_len:
            block_size = remaining
        else:
            block_size = random.randint(min_len, min(max_len, remaining))

        block = list(entries[i : i + block_size])
        blocks.append(block)
        i += block_size

    # Merge trailing block if it ended up shorter than min_len and there is a previous block
    if len(blocks) >= 2 and len(blocks[-1]) < min_len:
        logger.debug("Merging trailing short block of length %s", len(blocks[-1]))
        blocks[-2].extend(blocks[-1])
        blocks.pop()

    return blocks


def choose_tasks(task_options: Sequence[str]) -> List[str]:
    if not task_options:
        return []
    # Determine how many tasks to append (1 or 2 if available)
    max_tasks = min(len(task_options), 2)
    task_count = random.randint(1, max_tasks)
    return random.sample(task_options, k=task_count)


def format_appendix(tasks: Sequence[str]) -> str:
    if not tasks:
        return ""
    task_text = "; ".join(tasks)
    return f" - Úlohy: {task_text}"


def apply_block(tasks_for_position: Sequence[str], block: Sequence[DayEntry], ws: Worksheet) -> None:
    selected_tasks = choose_tasks(tasks_for_position)
    if not selected_tasks:
        logger.warning(
            "No tasks available for position in sheet %s; skipping rows %s-%s",
            ws.title,
            block[0].row_index,
            block[-1].row_index,
        )
        return

    appendix = format_appendix(selected_tasks)
    if not appendix:
        return

    for entry in block:
        cell = ws.cell(row=entry.row_index, column=COLUMN_ACTIVITY)
        cell.value = f"{entry.base_text}{appendix}"


def process_sheet(
    ws: Worksheet,
    task_map: Dict[str, List[str]],
    min_block: int,
    max_block: int,
    month_name: str,
    year: int,
) -> None:
    position_raw = ws.cell(row=12, column=COLUMN_ACTIVITY).value
    if not isinstance(position_raw, str) or not position_raw.strip():
        logger.info("Skipping sheet %s due to missing working position in E12", ws.title)
        return

    position = position_raw.strip()
    tasks_for_position = task_map.get(position)
    if not tasks_for_position:
        logger.info("No tasks defined for position '%s' (sheet %s); skipping", position, ws.title)
        return

    entries = extract_day_entries(ws)
    if not entries:
        logger.info("No eligible entries found in sheet %s", ws.title)
        return

    blocks = partition_entries(entries, min_len=min_block, max_len=max_block)
    for block in blocks:
        apply_block(tasks_for_position, block, ws)


def should_skip_sheet(sheet_name: str, skip_prefixes: Iterable[str]) -> bool:
    lowered = sheet_name.lower()
    return any(lowered.startswith(prefix.lower()) for prefix in skip_prefixes)


def ensure_output_path(path: Path, force: bool) -> None:
    if path.exists() and not force:
        raise FileExistsError(
            f"Output file {path} already exists. Use --force to overwrite or specify a different path."
        )
    path.parent.mkdir(parents=True, exist_ok=True)


def main() -> None:
    args = parse_args()

    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")

    if args.seed is not None:
        random.seed(args.seed)

    task_map = load_tasks(args.tasks_json)

    if not args.workbook.is_file():
        raise FileNotFoundError(f"Workbook not found: {args.workbook}")

    ensure_output_path(args.output, args.force)

    wb: Workbook = load_workbook(args.workbook)

    processed_sheets = 0
    for sheet_name in wb.sheetnames:
        if should_skip_sheet(sheet_name, args.skip_sheets):
            logger.info("Skipping sheet %s due to skip list", sheet_name)
            continue

        ws = wb[sheet_name]
        process_sheet(
            ws,
            task_map=task_map,
            min_block=args.min_block,
            max_block=args.max_block,
            month_name=args.month_name,
            year=args.year,
        )
        processed_sheets += 1

    wb.save(args.output)
    logger.info("Augmented workbook saved to %s", args.output)
    logger.info("Processed %s sheets", processed_sheets)


if __name__ == "__main__":
    main()
